#!/usr/bin/env python3
"""End-to-end smoke test: fill the workbook from the example spec and assert it computed.

Run from the package root:  python3 tests/test_smoke.py
Exits non-zero on failure. No API keys needed (uses the bundled template + example spec).
"""
import os, sys, json, tempfile, subprocess

PKG = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PKG)


def main():
    spec_path = os.path.join(PKG, "assets", "example.spec.json")
    assert os.path.exists(spec_path), "missing assets/example.spec.json"
    out = os.path.join(tempfile.gettempdir(), "smoke - STR COC Analysis.xlsx")

    r = subprocess.run([sys.executable, os.path.join(PKG, "fill_deal_analyzer.py"),
                        spec_path, "--out", out], capture_output=True, text=True)
    print(r.stdout[-800:])
    if r.returncode != 0:
        print(r.stderr[-800:]); sys.exit("fill_deal_analyzer failed")

    assert os.path.exists(out), "workbook not written"
    assert "CASH-ON-CASH" in r.stdout, "no headline metrics printed"
    assert "VERIFY:" in r.stdout, "verify step did not run"

    # confirm the Summary snapshot carries a real (non-blank) CoC value
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert "Summary (Pono)" in wb.sheetnames, "Summary sheet missing"
    assert "Pono Notes" in wb.sheetnames, "Pono Notes sheet missing"
    coc = wb["Summary (Pono)"]["B6"].value
    assert isinstance(coc, (int, float)), f"CoC snapshot not numeric: {coc!r}"

    print(f"\nSMOKE OK — workbook written, CoC snapshot = {coc:.1%}, both helper sheets present.")


if __name__ == "__main__":
    main()
