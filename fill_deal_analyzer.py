#!/usr/bin/env python3
"""
fill_deal_analyzer.py — fill the STR Cash-on-Cash workbook from a JSON spec
and print the computed headline metrics.

Why this exists: openpyxl writes input cells but does NOT recalculate the
workbook's formulas, and there is no LibreOffice on the mini to recalc headless.
So this script (a) writes every input cell from a spec, then (b) replicates the
STR Deal Analyzer formula chain in Python to print the results for the report.
The saved .xlsx still recalculates normally when opened in Excel/Numbers/Sheets.

Usage:
    python3 scripts/fill_deal_analyzer.py spec.json
    python3 scripts/fill_deal_analyzer.py spec.json --out "/path/Property - STR COC Analysis.xlsx"

The spec is a flat JSON object. Any key omitted keeps the template's default.
See references/str-deal-analyzer/README.md for the full key -> cell map and the
research checklist the /analyze-deal skill follows to gather these values.
"""
import json, sys, os, shutil, argparse

# key -> (cell). These are the INPUT (blue) cells on the 'STR Deal Analyzer' tab.
# Everything else on every tab is formula-driven and recalculates from these.
CELLS = {
    # property overview
    "address": "C5", "listing_link": "C6", "source_note": "C7",
    "market": "F6", "property_type": "F7",
    "beds": "C8", "baths": "F8", "sqft": "C9", "year_built": "F9",
    "lot_size": "C10", "amenities": "F10", "notes": "C11",
    # purchase
    "purchase_price": "C14", "arv": "F14", "down_pct": "C15",
    "closing_pct": "C16", "agent_comm": "C17", "transaction_coordinator": "F17",
    # rehab & setup
    "rehab": "C21", "setup_expenses": "F21", "furn_per_sqft": "C22",
    "furn_override": "C23", "legal": "C24", "permits": "F24", "photos": "C25",
    # financing
    "interest_rate": "F29", "term_years": "C30", "interest_only": "F30", "pmi_pct": "F31",
    # revenue
    "adr": "C36", "occupancy": "F36", "active_days": "C37", "gross_override": "F37",
    # channel mix
    "airbnb_pct": "C41", "vrbo_pct": "C42", "booking_pct": "C43", "direct_pct": "C44",
    "airbnb_fee": "F41", "vrbo_fee": "F42", "booking_fee": "F43", "processing_fee": "F44",
    # cleaning
    "turnovers_mo": "C48", "cost_per_turnover": "F48",
    # mgmt
    "mgmt_fee_pct": "C56",
    # fixed expenses
    "property_tax": "C60", "insurance": "F60", "hoa_annual": "C61",
    # variable expenses (monthly unless noted)
    "repairs_mo": "C65", "supplies_per_turn": "C66", "refunds_mo": "C67",
    "utilities_mo": "C68", "internet_mo": "C69", "landscaping_mo": "C70", "spa_mo": "C71",
    # tax
    "tax_bracket": "C85", "land_pct": "F85",
}

def fmt(x):
    return f"${x:,.0f}"

def compute(s, ws=None):
    """Replicate the model. s = dict of effective input values (spec merged over template)."""
    g = lambda k, d=0.0: s.get(k, d)
    price = g("purchase_price"); arv = g("arv") or price
    down = arv * g("down_pct"); closing = arv * g("closing_pct")
    cash_close = down + closing + g("agent_comm") + g("transaction_coordinator")
    loan = price - down
    rate = g("interest_rate"); term = g("term_years", 30)
    io = str(g("interest_only", "No")).strip().lower() == "yes"
    r = rate / 12; n = term * 12
    pi = (loan * r) if io else (loan * r / (1 - (1 + r) ** -n) if r and n else 0)
    pmi = loan * g("pmi_pct") / 12
    debt = (pi + pmi) * 12
    sqft = g("sqft")
    furn = g("furn_override") if g("furn_override") > 0 else sqft * g("furn_per_sqft")
    setup = g("rehab") + g("setup_expenses") + furn + g("legal") + g("permits") + g("photos")
    total_cash = cash_close + setup
    # revenue. DEFINITION: gross = total accommodation revenue INCLUDING cleaning fees
    # (this is what AirROI's ttm_revenue reports). The waterfall then subtracts the
    # cleaning payout, so cleaning nets to ~0 when it's pass-through, or leaves the real
    # markup/shortfall. 'occupancy' here is CALENDAR occupancy (booked nights / 365), NOT
    # AirROI's adjusted occupancy (booked / available) — feeding adjusted occ here would
    # overstate revenue. The ADR x days x occ fallback is nightly-only, so we add the
    # cleaning payout back to keep the same "includes cleaning" definition as gross_override.
    days = g("active_days", 365)
    turns = g("turnovers_mo")
    cleaning = turns * g("cost_per_turnover") * 12
    gross = g("gross_override") if g("gross_override") > 0 else g("adr") * days * g("occupancy") + cleaning
    blended = (g("airbnb_pct") * g("airbnb_fee") + g("vrbo_pct") * g("vrbo_fee")
               + g("booking_pct") * g("booking_fee") + g("direct_pct") * g("processing_fee"))
    ota = gross * blended
    host = gross - ota - cleaning
    mgmt = host * g("mgmt_fee_pct")
    net = host - mgmt
    fixed = g("property_tax") + g("insurance") + g("hoa_annual")
    variable = (g("repairs_mo") * 12 + g("supplies_per_turn") * turns * 12 + g("refunds_mo") * 12
                + g("utilities_mo") * 12 + g("internet_mo") * 12 + g("landscaping_mo") * 12 + g("spa_mo") * 12)
    noi = net - fixed - variable
    cf = noi - debt
    coc = cf / total_cash if total_cash else 0
    cap = noi / price if price else 0
    dscr = noi / debt if debt else 0
    return dict(down=down, closing=closing, cash_close=cash_close, loan=loan, pi=pi, debt=debt,
                furn=furn, setup=setup, total_cash=total_cash, gross=gross, ota=ota, cleaning=cleaning,
                host=host, mgmt=mgmt, net=net, fixed=fixed, variable=variable, noi=noi, cf=cf,
                coc=coc, cap=cap, dscr=dscr)

def _write_summary(wb, spec, eff, m):
    """Bake a computed-values snapshot as a front sheet of LITERAL numbers.

    The model tabs are formulas with no cached results, so they read blank in any viewer
    that doesn't recalc (Quick Look, sometimes Numbers). This sheet holds real values that
    show everywhere; the live tabs still recalc on open for what-if editing.
    """
    import datetime
    name = "Summary (Pono)"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    USD = '"$"#,##0'
    PCT1 = '0.0%'
    PCT2 = '0.00%'
    rows = [
        ("STR Cash-on-Cash — Computed Snapshot", None, None),
        (spec.get("address", ""), None, None),
        (f"List/price: ${spec.get('purchase_price', 0):,.0f}   |   filled {datetime.date.today()}", None, None),
        (None, None, None),
        ("HEADLINE RETURNS", None, None),
        ("Cash-on-Cash (pre-tax)", m["coc"], PCT1),
        ("Cap rate", m["cap"], PCT2),
        ("DSCR", round(m["dscr"], 2), '0.00'),
        ("Monthly cash flow", m["cf"] / 12, USD),
        ("Annual cash flow", m["cf"], USD),
        (None, None, None),
        ("CASH NEEDED", None, None),
        ("Down payment", m["down"], USD),
        ("Cash to close", m["cash_close"], USD),
        ("Rehab", eff.get("rehab", 0), USD),
        ("Furnishings", m["furn"], USD),
        ("TOTAL CASH NEEDED", m["total_cash"], USD),
        (None, None, None),
        ("FINANCING", None, None),
        ("Loan amount", m["loan"], USD),
        ("Monthly P&I", m["pi"], USD),
        (None, None, None),
        ("REVENUE & NOI", None, None),
        ("Gross revenue", m["gross"], USD),
        ("Net revenue to owner", m["net"], USD),
        ("NOI", m["noi"], USD),
        ("Annual management fee (StayAscend)", m["mgmt"], USD),
        (None, None, None),
        ("Snapshot computed by Pono when the file was filled. The other tabs are the live",
         None, None),
        ("model and recalculate when opened in Excel or Numbers (edit inputs to what-if).",
         None, None),
    ]
    from openpyxl.styles import Font
    for i, (label, val, fmt_) in enumerate(rows, 1):
        ws[f"A{i}"] = label
        if i == 1:
            ws[f"A{i}"].font = Font(bold=True, size=13)
        elif label in ("HEADLINE RETURNS", "CASH NEEDED", "FINANCING", "REVENUE & NOI"):
            ws[f"A{i}"].font = Font(bold=True)
        if val is not None:
            c = ws[f"B{i}"]
            c.value = val
            if fmt_:
                c.number_format = fmt_
            if label in ("Cash-on-Cash (pre-tax)", "TOTAL CASH NEEDED"):
                c.font = Font(bold=True)
    return ws


def _write_reasoning(wb, spec, eff, m):
    """Append a 'Pono Notes' sheet at the END: the analyst's thought process behind the inputs.

    Driven by the spec's `reasoning` list (the skill fills this as it underwrites). Each entry
    may be a {"topic","detail"} dict, a [topic, detail] pair, or a plain string (full-width
    bullet). If `reasoning` is absent, it synthesizes a minimal log from source_note + notes so
    the sheet is never empty. Anyone wondering 'why did Pono enter this?' reads it here.
    """
    import datetime
    from openpyxl.styles import Font, Alignment
    name = "Pono Notes"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)                      # no index → appended last
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    top = Font(bold=True, size=13)
    hdr = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def normalize(entries):
        out = []
        for e in entries:
            if isinstance(e, dict):
                out.append((str(e.get("topic", "")), str(e.get("detail", ""))))
            elif isinstance(e, (list, tuple)):
                out.append((str(e[0]) if len(e) > 0 else "", str(e[1]) if len(e) > 1 else ""))
            else:
                out.append(("", str(e)))
        return out

    rows = normalize(spec.get("reasoning", []))
    if not rows:                                    # fallback: synthesize from existing free-text
        if spec.get("source_note"):
            rows.append(("Revenue basis", str(spec["source_note"])))
        if spec.get("notes"):
            rows.append(("Property notes", str(spec["notes"])))
        if not rows:
            rows.append(("", "No analyst reasoning was recorded for this run."))

    # header block
    ws["A1"] = "Pono Notes — analyst reasoning & update log"
    ws["A1"].font = top
    ws["A2"] = spec.get("address", "")
    ws["A3"] = (f"Filled {datetime.date.today()}   |   CoC {m['coc']:.1%}   |   "
                f"gross ${m['gross']:,.0f}   |   total cash ${m['total_cash']:,.0f}")
    ws["A4"] = "Why each input is what it is. The numbers live on the other tabs; the why lives here."
    ws["A4"].font = Font(italic=True)

    r = 6
    ws[f"A{r}"] = "TOPIC"; ws[f"B{r}"] = "REASONING"
    ws[f"A{r}"].font = hdr; ws[f"B{r}"].font = hdr
    r += 1
    for topic, detail in rows:
        ws[f"A{r}"] = topic
        ws[f"A{r}"].font = hdr
        ws[f"A{r}"].alignment = wrap
        ws[f"B{r}"] = detail
        ws[f"B{r}"].alignment = wrap
        r += 1
    return ws


def _excel_recalc(out):
    """macOS + Microsoft Excel only: open the saved workbook, force a recalc (touch an input
    cell so the workbook is dirty), and save — so the LIVE formula tabs carry cached results and
    display in ANY viewer (Quick Look / Numbers / Sheets), not just after a manual recalc.

    Why this is needed: openpyxl writes formulas but no cached values, and fullCalcOnLoad alone
    doesn't persist a recalc unless something is marked dirty. Real Excel handles PMT etc. that
    pycel can't. No-op (returns False) when not on a Mac or Excel/osascript isn't available.
    """
    import platform, subprocess
    if platform.system() != "Darwin":
        return False
    script = '''
set f to POSIX file "%s" as alias
tell application "Microsoft Excel"
    open f
    set wb to active workbook
    set sh to worksheet "STR Deal Analyzer" of wb
    set value of cell "C5" of sh to (get value of cell "C5" of sh)
    close wb saving yes
end tell
''' % out
    try:
        r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=180)
        return r.returncode == 0
    except Exception:
        return False


def _verify(out, spec, m, recalced=False):
    """Re-open the saved file and confirm every input landed and the snapshot is populated.

    Catches the 'blank workbook' class of bug before we report success: missing inputs,
    empty summary cells, or a CoC that didn't make it in. When a live Excel recalc ran, also
    confirm the live formula tab actually carries cached values (not just the snapshot).
    """
    import openpyxl
    problems = []
    if recalced:
        live = openpyxl.load_workbook(out, data_only=True)["STR Deal Analyzer"]
        if not isinstance(live["C106"].value, (int, float)):
            problems.append("live recalc ran but CoC (C106) cell still blank")
    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["STR Deal Analyzer"]
    for k, v in spec.items():
        if k in CELLS and ws[CELLS[k]].value in (None, ""):
            problems.append(f"input '{k}' ({CELLS[k]}) did not write")
    if "Summary (Pono)" not in wb.sheetnames:
        problems.append("summary sheet missing")
    else:
        s = wb["Summary (Pono)"]
        for cell, what in [("B6", "CoC"), ("B8", "DSCR"), ("B17", "total cash"),
                           ("B24", "gross revenue")]:
            if s[cell].value in (None, ""):
                problems.append(f"summary {what} ({cell}) blank")
    if not wb.calculation.fullCalcOnLoad:
        problems.append("fullCalcOnLoad not set (live tabs won't recalc on open)")
    if problems:
        print("VERIFY: ⚠ issues found —")
        for p in problems:
            print("   -", p)
    else:
        print(f"VERIFY: ✓ inputs written, snapshot populated (CoC {m['coc']:.1%}), "
              f"recalc-on-open set.")
    return not problems


def _default_template():
    # config.TEMPLATE if the user made a config.py, else the bundled assets/ template.
    try:
        import config
        return config.TEMPLATE
    except Exception:
        return os.path.join(os.path.dirname(__file__), "assets", "template.xlsx")


def template_defaults(template=None):
    """The numeric default for every input cell, read from the blank template.

    This is the SAME default set the fill path overlays the spec onto, so any caller that
    builds an effective spec (e.g. deal_probabilistic) matches the workbook instead of
    inheriting some other property's full spec. Keys the spec omits keep these values.
    """
    import openpyxl
    ws = openpyxl.load_workbook(template or _default_template(), data_only=False)["STR Deal Analyzer"]
    eff = {}
    for k, cell in CELLS.items():
        v = ws[cell].value
        if isinstance(v, (int, float)):
            eff[k] = v
    return eff


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("spec")
    ap.add_argument("--template", default=_default_template())
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    import openpyxl

    spec = json.load(open(args.spec))
    out = args.out or os.path.join(os.path.expanduser("~/Downloads"),
                                   f"{spec.get('address','Property')} - STR COC Analysis.xlsx")
    shutil.copyfile(os.path.abspath(args.template), out)
    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["STR Deal Analyzer"]

    # read template defaults for any numeric input the spec omits (so compute() matches the sheet)
    eff = template_defaults(args.template)
    eff.update({k: v for k, v in spec.items() if k in CELLS})

    # write spec values into the sheet
    written = 0
    for k, v in spec.items():
        if k in CELLS:
            ws[CELLS[k]] = v
            written += 1
    m = compute(eff)
    # openpyxl writes formulas but leaves their cached results empty, so the live tabs read
    # blank in any viewer that shows the stored value (Quick Look, Numbers). Bake a snapshot
    # sheet of literal values that shows everywhere, AND force the live tabs to recalc on open.
    _write_summary(wb, spec, eff, m)
    _write_reasoning(wb, spec, eff, m)
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    recalced = _excel_recalc(out)            # bake cached values into the live tabs (Mac+Excel)
    _verify(out, spec, m, recalced)
    print(f"Live tabs recalculated via Excel: {'yes' if recalced else 'no (snapshot sheet still populated)'}")
    print(f"Saved: {out}")
    print(f"Wrote {written} input cells.\n")
    print("=== CASH NEEDED ===")
    print("Down payment:        ", fmt(m["down"]))
    print("Cash to close:       ", fmt(m["cash_close"]))
    print("Furnishings:         ", fmt(m["furn"]))
    print("TOTAL CASH NEEDED:   ", fmt(m["total_cash"]))
    print("\n=== FINANCING ===")
    print("Loan amount:         ", fmt(m["loan"]))
    print("Monthly P&I:         ", fmt(m["pi"]))
    print("Annual debt service: ", fmt(m["debt"]))
    print("\n=== REVENUE ===")
    print("Gross revenue:       ", fmt(m["gross"]))
    print("Net rev to owner:    ", fmt(m["net"]))
    print("\n=== RETURNS ===")
    print("NOI:                 ", fmt(m["noi"]))
    print("Pre-tax cash flow:   ", fmt(m["cf"]), f"({fmt(m['cf']/12)}/mo)")
    print("CASH-ON-CASH:        ", f"{m['coc']:.1%}")
    print("Cap rate:            ", f"{m['cap']:.2%}")
    print("DSCR:                ", f"{m['dscr']:.2f}")
    # --- B6: manager (StayAscend) economics — what YOU earn managing it ---
    mgmt = m["mgmt"]
    five_yr = sum(mgmt * (1.03 ** y) for y in range(5))   # 3% revenue growth
    print("\n=== MANAGER ECONOMICS (StayAscend) ===")
    print("Annual management fee:", fmt(mgmt))
    print("5-yr mgmt revenue:   ", fmt(five_yr), "(3% growth)")
    print("Mgmt fee / gross:    ", f"{(mgmt / m['gross']):.1%}" if m["gross"] else "n/a")

    # --- B5 + B7 + tax-tier: context & checks ---
    notes = []
    pt = str(spec.get("property_type", "")).lower()
    if "condo" in pt or "townhome" in pt:
        notes.append("FINANCING: condo/townhome — verify it's WARRANTABLE. High investor/STR "
                     "concentration (common in resort condos) can block conventional loans, "
                     "forcing a DSCR/portfolio loan at a higher rate + bigger down payment than "
                     "the 25%/7% assumed here.")
    mkt = spec.get("permit_market")
    if mkt:
        try:
            import permit_costs
            if mkt in permit_costs.MARKETS:
                M = permit_costs.MARKETS[mkt]
                w = permit_costs.tax_tier_warning(mkt, spec.get("property_tax"), spec.get("purchase_price"))
                if w:
                    notes.append("TAX TIER: " + w)
                t = M.get("taxes", {})
                rate = t.get("get_pct", 0) + t.get("state_tat_pct", 0) + t.get("county_tat_pct", 0)
                if rate:
                    notes.append(f"LODGING TAX: ~{rate:.1%} (GET+TAT) is collected from guests on "
                                 f"${m['gross']:,.0f} gross = ~{fmt(m['gross'] * rate)}/yr remitted. "
                                 f"Pass-through (not owner cost), but it raises the guest's all-in "
                                 f"price, which pressures the occupancy assumed above.")
                cf = M.get("cost_factors", {})
                sug = cf.get("furnishing_per_sqft")
                fps = spec.get("furn_per_sqft", eff.get("furn_per_sqft"))
                if sug and fps and fps < sug * 0.9:
                    notes.append(f"COST: furnishing at ${fps}/sqft looks low for this market "
                                 f"(~${sug}/sqft with island shipping/labor). {cf.get('insurance_note','')}")
        except Exception:
            pass
    if notes:
        print("\n=== CHECKS & CONTEXT ===")
        for nidx, n in enumerate(notes, 1):
            print(f"[{nidx}] {n}")


if __name__ == "__main__":
    main()
